import os
import numpy as np
import re
import openai
import pandas as pd
import base64
import streamlit as st
import openpyxl
#import anthropic
import pickle
import tiktoken
from sklearn.neighbors import NearestNeighbors
from joblib import dump, load
#from transformers import GPT2TokenizerFast
from scipy.spatial import distance_matrix
#from sentence_transformers import SentenceTransformer
from nltk.cluster.util import cosine_distance
import constants as C
from openpyxl import Workbook
wb = Workbook()

def salva_q(value):
    path = C.directory + 'Archivio/Elenco_risultati.xlsx'
    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    #data = (player, int(value), datetime)
    ws.append(value)

    wb.save(C.directory + 'Archivio/Elenco_risultati.xlsx')

def save_partita(data1, path):
    #res = ' '.join(str(val) for val in data1)
    #string_list1 = res.split(",")
    #string_list2 = data2.split(",")
    embedded = pd.DataFrame(data=data1)
    fp = open(path, 'w')
    embedded.to_csv(path, sep=',', na_rep='', float_format=int, header=True, index=False, index_label=None,)
